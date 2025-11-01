import streamlit as st
import pandas as pd
import random
import io
from datetime import datetime, timedelta, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =============================
# 🎨 DARK DISPATCH MODE STYLE
# =============================
st.set_page_config(page_title="SMKL Scheduling Assistant — Dark Mode", layout="centered")

st.markdown("""
<style>
    .stApp, .main { background-color: #1e1e1e; color: #e5e5e5; }
    h1, h2, h3, h4 { color: #f5f5f5 !important; }
    .block-container { padding-top: 1.25rem; padding-bottom: 1.25rem; }
    div[data-testid="stExpander"] div[role="button"] p { color: #ccc !important; }
    .panel { padding: 8px 10px; border-radius: 8px; color: #fff; margin: 6px 0; }
</style>
""", unsafe_allow_html=True)

# =============================
# ⚙️ HELPERS
# =============================
DAY_NAMES = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

LABEL_TEXT = {
    "DOT": "DOT Route",
    "DOT-HelperRoute": "DOT-HelperRoute",
    "DOT-Helper": "DOT-Helper",
    "XL": "XL",
    "Standby": "Standby",
}

def scheduled_on_day(driver, day):
    v = driver["days"].get(day)
    if pd.isna(v): return False
    s = str(v).strip().lower()
    return s in ("1", "dot")

def infer_dot_cert(driver):
    for v in driver["days"].values():
        if pd.isna(v): continue
        if str(v).strip().lower() == "dot":
            return True
    return False

def week_start_from_number(year:int, week:int):
    # ISO week Monday -> minus 1 day to get Sunday
    return (datetime.fromisocalendar(year, week, 1).date() - timedelta(days=1))

def safe_sheet_name(name: str) -> str:
    invalid = '[]:*?/\\'
    cleaned = ''.join('_' if c in invalid else c for c in name)[:31]
    return cleaned.strip()

def write_df_to_sheet_openpyxl(wb, sheet_name: str, df: pd.DataFrame):
    # Delete if exists, then create
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    # write headers
    headers = list(df.columns)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    # write rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    # simple autosize
    for i, col in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(col)) + 2)

# =============================
# 🖥️ APP HEADER
# =============================
st.title("🚚 SMKL Scheduling Assistant — Designed by Timmy Nguyen 😎")
st.caption("Plan smart. Drive safe. Rest easy.")

uploaded = st.file_uploader("📄 Upload weekly Excel (rows 14–90, cols D–L):", type=["xlsx"])

if uploaded:
    # Keep a copy of the raw bytes so we can reopen with openpyxl later
    original_bytes = uploaded.getvalue()

    # Parse driver grid with pandas
    df_grid = pd.read_excel(io.BytesIO(original_bytes), header=None, skiprows=13, nrows=77)
    st.success("✅ File uploaded successfully!")

    # Build driver objects
    drivers = []
    for _, row in df_grid.iterrows():
        first = str(row[3]).strip() if not pd.isna(row[3]) else ""
        last = str(row[4]).strip() if not pd.isna(row[4]) else ""
        if not first and not last:
            continue
        days = {DAY_NAMES[i]: row[5 + i] for i in range(7)}
        drivers.append({"name": f"{first} {last}".strip(), "days": days})
    dot_map = {d["name"]: infer_dot_cert(d) for d in drivers}

    # Inputs
    year_now = datetime.today().year
    week_input = st.number_input("📅 Enter week number (current year):", min_value=1, max_value=53, step=1, value=1)
    wk_start = week_start_from_number(year_now, int(week_input))
    st.write(f"🗓 Week {int(week_input)} starts **Sunday {wk_start}**")

    new_list = [s.strip() for s in st.text_area("🆕 Paste NEW drivers (one per line):").splitlines() if s.strip()]
    semi = [s.strip() for s in st.text_area("🚫 Paste SEMI-restricted drivers (cannot do DOT / DOT-HelperRoute):").splitlines() if s.strip()]

    st.divider()
    st.subheader("📦 Enter Daily Route Counts")

    day_inputs = {}
    for idx, day in enumerate(DAY_NAMES):
        with st.expander(f"🗓 {day} ({(wk_start + timedelta(days=idx)).strftime('%m/%d')})", expanded=False):
            dot_hr = st.number_input(f"{day} — DOT-HelperRoute", min_value=0, step=1, key=f"{day}_hr")
            dot_h  = st.number_input(f"{day} — DOT-Helpers",     min_value=0, step=1, key=f"{day}_h")
            dot_r  = st.number_input(f"{day} — DOT Routes",       min_value=0, step=1, key=f"{day}_r")
            xl     = st.number_input(f"{day} — XL Routes",        min_value=0, step=1, key=f"{day}_xl")
            day_inputs[day] = {"dot_helperroute": dot_hr, "dot_helper": dot_h, "dot": dot_r, "xl": xl}

    if st.button("🚀 Generate Schedule"):
        progress = st.progress(0.0)

        # Trackers
        dot_weekly_count = {n: 0 for n, is_dot in dot_map.items() if is_dot}
        dot_stepvan_count = {n: 0 for n, is_dot in dot_map.items() if is_dot}
        standby_tracker = {}
        all_day_dataframes = {}
        # For updating the weekly sheet later
        day_label_maps = {day: {} for day in DAY_NAMES}

        # ===== Day-by-day generation =====
        for idx, (day, counts) in enumerate(day_inputs.items()):
            assigned = {"DOT": [], "DOT-HelperRoute": [], "DOT-Helper": [], "XL": [], "Standby": []}
            scheduled_today = [d["name"] for d in drivers if scheduled_on_day(d, day)]

            # DOT candidates (respect semi restriction)
            dot_avail = [n for n in scheduled_today if dot_map.get(n, False) and n not in semi]
            eligible_dot = [n for n in dot_avail if dot_weekly_count.get(n, 0) < 2]

            # Prioritize DOT drivers with fewer step-van counts (fairness)
            random.shuffle(eligible_dot)
            eligible_dot.sort(key=lambda n: dot_stepvan_count.get(n, 0))  # 0 first, then 1

            # DOT routes
            assigned["DOT"] = eligible_dot[:counts["dot"]]
            for n in assigned["DOT"]:
                dot_weekly_count[n] += 1
                dot_stepvan_count[n] += 1

            # DOT-HelperRoute
            remaining_dot = [n for n in eligible_dot if n not in assigned["DOT"]]
            remaining_dot.sort(key=lambda n: dot_stepvan_count.get(n, 0))
            assigned["DOT-HelperRoute"] = remaining_dot[:counts["dot_helperroute"]]
            for n in assigned["DOT-HelperRoute"]:
                dot_weekly_count[n] += 1
                dot_stepvan_count[n] += 1

            # DOT-Helper (does NOT count as step-van, but counts toward weekly DOT cap)
            helper_pool = [n for n in scheduled_today if n not in assigned["DOT"] + assigned["DOT-HelperRoute"]]
            # Prefer DOT helpers who have fewer total DOT-counted assignments
            helper_pool.sort(key=lambda n: (0 if dot_map.get(n, False) else 1, dot_weekly_count.get(n, 0)))
            assigned["DOT-Helper"] = helper_pool[:counts["dot_helper"]]
            for n in assigned["DOT-Helper"]:
                if dot_map.get(n, False):
                    dot_weekly_count[n] += 1

            # XL: new drivers FIRST (but only if scheduled), then others
            new_sched = [n for n in new_list if n in scheduled_today]
            take_new = min(len(new_sched), counts["xl"])
            assigned["XL"] = new_sched[:take_new]
            need_xl = counts["xl"] - len(assigned["XL"])
            if need_xl > 0:
                rem = [n for n in scheduled_today if n not in assigned["DOT"] + assigned["DOT-HelperRoute"] +
                       assigned["DOT-Helper"] + assigned["XL"]]
                random.shuffle(rem)
                assigned["XL"].extend(rem[:need_xl])

            # Standby (cap 2 per week)
            standby_pool = [n for n in scheduled_today if n not in set(
                assigned["DOT"] + assigned["DOT-HelperRoute"] + assigned["DOT-Helper"] + assigned["XL"]
            )]
            standby_final = []
            for s in standby_pool:
                if standby_tracker.get(s, 0) < 2:
                    standby_final.append(s)
                    standby_tracker[s] = standby_tracker.get(s, 0) + 1
            assigned["Standby"] = standby_final

            # ---- Build daily DataFrame
            rows = []
            for grp in ["DOT", "DOT-HelperRoute", "DOT-Helper", "XL", "Standby"]:
                rows.append({"Group": f"{grp} ({len(assigned[grp])})", "Driver": ""})
                for n in assigned[grp]:
                    rows.append({"Group": "", "Driver": n})
                rows.append({"Group": "", "Driver": ""})
            df_day = pd.DataFrame(rows)
            all_day_dataframes[day] = df_day

            # ---- Build label map for weekly update
            for grp in ["DOT", "DOT-HelperRoute", "DOT-Helper", "XL", "Standby"]:
                for n in assigned[grp]:
                    day_label_maps[day][n] = grp  # raw labels; convert to text later

            # ---- Panels (flat, colored)
            st.markdown(f"### {day} {(wk_start + timedelta(days=idx)).strftime('%m/%d')}")
            for grp, color, emoji in [
                ("DOT", "#22c55e", "🚛"),
                ("DOT-HelperRoute", "#0ea5e9", "🚐"),
                ("DOT-Helper", "#60a5fa", "🧑‍🤝‍🧑"),
                ("XL", "#eab308", "📦"),
                ("Standby", "#9ca3af", "💤"),
            ]:
                st.markdown(f"<div class='panel' style='background:{color};'><b>{emoji} {grp} ({len(assigned[grp])})</b></div>", unsafe_allow_html=True)
                if assigned[grp]:
                    st.write(", ".join(assigned[grp]))
            progress.progress((idx + 1) / 7.0)

        # ===== Update the ORIGINAL workbook in memory =====
        st.divider()
        st.info("🔄 Updating original weekly sheet (replacing 1/DOT with assignments)…")

        wb = load_workbook(io.BytesIO(original_bytes))
        ws = wb[wb.sheetnames[0]]  # weekly grid

        # Weekly grid: rows 14–90, Col D=4 (first), E=5 (last), days F..L (6..12)
        for r in range(14, 91):
            first = ws.cell(r, 4).value
            last  = ws.cell(r, 5).value
            if not first and not last:
                continue
            name = f"{str(first).strip()} {str(last).strip()}".strip()
            for day_idx, col in enumerate(range(6, 13)):
                day_name = DAY_NAMES[day_idx]
                cell = ws.cell(r, col)
                v = cell.value
                s = str(v).strip().lower() if v is not None else ""
                # Always check if this driver has any assignment for that day
                label = day_label_maps.get(day_name, {}).get(name)

                # If driver is scheduled for that day (any assignment), update label
                if label:
                    cell.value = LABEL_TEXT.get(label, label)
                # Otherwise, leave their cell untouched (don’t erase anything)

        st.success("✅ Weekly sheet updated (colors preserved).")

        # ===== Write/replace per-day sheets in same workbook =====
        st.info("🗓 Writing per-day sheets (Sun–Sat)…")
        for idx, (day, df_day) in enumerate(all_day_dataframes.items()):
            day_title = f"{day} {(wk_start + timedelta(days=idx)).strftime('%m_%d')}"
            write_df_to_sheet_openpyxl(wb, safe_sheet_name(day_title), df_day)
        st.success("✅ Daily sheets written.")

        # No Weekly_Summary tab — intentionally removed
        # Build final download
        out_buf = io.BytesIO()
        wb.save(out_buf)

        st.success("🟩 Step-van fairness applied | 🟩 Standby cap complete | 🟩 All sheets synchronized")
        st.download_button(
            "📥 Download Updated Workbook",
            data=out_buf.getvalue(),
            file_name=f"SMKL_schedule_week_{int(week_input)}_updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("👆 Upload your Excel file to start.")

